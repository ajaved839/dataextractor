from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from xlxs_dataextractor.exporter import write_results
from xlxs_dataextractor.extractor import extract_workbook


def test_extracts_discharge_summary_fields(tmp_path: Path) -> None:
    sample = tmp_path / "sample.xlsx"
    _create_sample_workbook(sample)

    result = extract_workbook(sample)

    assert result.data["MR #"] == "OHJT13362126"
    assert result.data["Patient Name"] == "MISS FAIZA KHADIM"
    assert result.data["Age"] == "41"
    assert result.data["Sex"] == "Female"
    assert result.data["Contact"] == "3217676995"
    assert result.data["Date of Admission"] == "30/5/2026"
    assert result.data["Date of Operation"] == "30/5/2026"
    assert result.data["Date of Discharge"] == "31/5/2026"
    assert result.data["Diagnosis"] == "Morbid Obesity"
    assert "Sleeve Gastrectomy" in result.data["Procedure"]
    assert "Gastrolysis done" in result.data["Operative Notes"]
    assert "Stitches removal after 9 days" in result.data["Follow Up Instructions"]
    assert "Inj R/L 500mg IV" in result.data["Discharge Medications"]


def test_writes_combined_output(tmp_path: Path) -> None:
    sample = tmp_path / "sample.xlsx"
    output = tmp_path / "output.xlsx"
    _create_sample_workbook(sample)

    result = extract_workbook(sample)
    write_results([result], output)

    workbook = load_workbook(output)
    worksheet = workbook.active

    assert worksheet["A1"].value == "Source File"
    assert worksheet["B2"].value == "OHJT13362126"
    assert worksheet["C2"].value == "MISS FAIZA KHADIM"


def test_extracts_contact_with_merged_labels(tmp_path: Path) -> None:
    sample = tmp_path / "merged_contact.xlsx"
    _create_screenshot_layout_workbook(sample)

    result = extract_workbook(sample)

    assert result.data["Contact"] == "3217676995"
    assert result.data["Date of Admission"] == "30/5/2026"
    assert result.data["Patient Name"] == "adeel avedasm"


def test_appends_to_existing_extract(tmp_path: Path) -> None:
    first = tmp_path / "first.xlsx"
    second = tmp_path / "second.xlsx"
    existing = tmp_path / "existing.xlsx"
    output = tmp_path / "output.xlsx"

    _create_sample_workbook(first)
    _create_screenshot_layout_workbook(second)

    first_result = extract_workbook(first)
    write_results([first_result], existing)

    second_result = extract_workbook(second)
    write_results([second_result], output, existing_path=existing)

    workbook = load_workbook(output)
    worksheet = workbook.active

    assert worksheet.max_row == 3
    assert worksheet["C2"].value == "MISS FAIZA KHADIM"
    assert worksheet["C3"].value == "adeel avedasm"
    assert worksheet["F3"].value == "3217676995"


def _create_screenshot_layout_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active

    worksheet["A1"] = "DISCHARGE SUMMARY"
    worksheet["A7"] = "MR #"
    worksheet["B7"] = "OHJT13362126"
    worksheet["C7"] = "AGE"
    worksheet["D7"] = "41 YEARS"
    worksheet["F7"] = "SEX"
    worksheet["G7"] = "male"
    worksheet["A9"] = "PATIENT NAME:"
    worksheet["B9"] = "adeel avedasm"
    worksheet.merge_cells("F9:G9")
    worksheet["F9"] = "CONTACT:"
    worksheet["H9"] = "3217676995"
    worksheet["A11"] = "DATE OF ADMISSION"
    worksheet["D11"] = "DATE OF OPERATION"
    worksheet["F11"] = "DATE OF DISCHARGE"
    worksheet["A12"] = "30/5/2026"
    worksheet["D12"] = "30/5/2026"
    worksheet.merge_cells("F12:G12")
    worksheet["F12"] = "31/5/2026"
    worksheet["A14"] = "PRESENTING COMPLAINTS"
    worksheet["A16"] = "Morbid Obesity"
    worksheet["A17"] = "ON LAB"
    worksheet["A19"] = "Hb 12.5 platelets 280,000/mm3"

    workbook.save(path)


def _create_sample_workbook(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active

    worksheet["A1"] = "DISCHARGE SUMMARY"
    worksheet["A2"] = "MR #"
    worksheet["B2"] = "OHJT13362126"
    worksheet["C2"] = "AGE"
    worksheet["D2"] = 41
    worksheet["E2"] = "YEARS"
    worksheet["F2"] = "SEX"
    worksheet["G2"] = "Female"
    worksheet["A3"] = "PATIENT NAME:"
    worksheet["B3"] = "MISS FAIZA KHADIM"
    worksheet["F3"] = "CONTACT:"
    worksheet["G3"] = "3217676995"
    worksheet["A4"] = "DATE OF ADMISSION"
    worksheet["A5"] = "30/5/2026"
    worksheet["D4"] = "DATE OF OPERATION"
    worksheet["D5"] = "30/5/2026"
    worksheet["F4"] = "DATE OF DISCHARGE"
    worksheet["F5"] = "31/5/2026"

    rows = [
        ("PRESENTING COMPLAINTS", "Morbid Obesity"),
        ("ON LAB", "Hb 12.5 platelets 280,000/mm3 HBsAg -ive, Anti HCV -ive, RFTs Normal"),
        ("DIAGNOSIS", "Morbid Obesity"),
        ("PROCEDURE", "Lap Sleeve Gastrectomy under GA with ETT"),
        (
            "HOSPITAL COURSE",
            "Patient was admitted in Private room. Laparoscopic Sleeve Gastrectomy was done.",
        ),
        (
            "OPERATIVE NOTES",
            "UAM Pt painted and drapped 5-Port Approach made. Gastrolysis done from the Greater Curvature.",
        ),
        (
            "FOLLOW UP INSTRUCTIONS",
            "Diet and follow up visit according to attach post operative instruction. Stitches removal after 9 days.",
        ),
        ("FOLLOW UP TIME & DATE", "After 9 days on 12/05/2026 on 10:00am"),
        (
            "DISCHARGE MEDICATIONS",
            "Inj R/L 500mg IV Once a day for 5 days Inj Tanzo 4.5g IV Twice a day for 5 days",
        ),
    ]

    row_number = 7
    for label, value in rows:
        worksheet.cell(row_number, 1, label)
        worksheet.cell(row_number + 1, 1, value)
        row_number += 2

    workbook.save(path)
