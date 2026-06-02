from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .fields import FIELD_COLUMNS, LABEL_ALIASES, SECTION_FIELDS


@dataclass(frozen=True)
class ExtractionResult:
    source_file: str
    data: dict[str, str]
    missing_fields: list[str]


@dataclass(frozen=True)
class CellText:
    row: int
    col: int
    value: str


def extract_workbook(path: str | Path) -> ExtractionResult:
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    worksheet = workbook.active

    cells = _read_text_cells(worksheet)
    rows = _rows_from_cells(cells)
    merged_lookup = _merged_value_lookup(worksheet)

    data = {column: "" for column in FIELD_COLUMNS}
    data["Source File"] = workbook_path.name

    for field in FIELD_COLUMNS:
        if field == "Source File":
            continue
        if field in SECTION_FIELDS:
            data[field] = _extract_section(field, rows)
        else:
            data[field] = _extract_inline(field, worksheet, cells, merged_lookup)

    missing = [field for field in FIELD_COLUMNS if field != "Source File" and not data[field]]
    return ExtractionResult(workbook_path.name, data, missing)


def _read_text_cells(worksheet: Worksheet) -> list[CellText]:
    cells: list[CellText] = []
    for row in worksheet.iter_rows():
        for cell in row:
            value = _clean_value(cell.value)
            if value:
                cells.append(CellText(cell.row, cell.column, value))
    return cells


def _rows_from_cells(cells: list[CellText]) -> dict[int, list[CellText]]:
    rows: dict[int, list[CellText]] = {}
    for cell in cells:
        rows.setdefault(cell.row, []).append(cell)
    for row_cells in rows.values():
        row_cells.sort(key=lambda item: item.col)
    return dict(sorted(rows.items()))


def _merged_value_lookup(worksheet: Worksheet) -> dict[tuple[int, int], str]:
    lookup: dict[tuple[int, int], str] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merged_range.min_row, merged_range.min_col)
        value = _clean_value(top_left.value)
        if not value:
            continue
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                lookup[(row, col)] = value
    return lookup


def _extract_inline(
    field: str,
    worksheet: Worksheet,
    cells: list[CellText],
    merged_lookup: dict[tuple[int, int], str],
) -> str:
    for cell in cells:
        if not _label_match(field, cell.value):
            continue

        after_label = _value_after_label(field, cell.value)
        if after_label and _is_valid_inline_value(field, after_label):
            return after_label

        right_value = _first_value_to_right(worksheet, cell.row, cell.col, merged_lookup, field)
        if right_value:
            return right_value

        if field == "Contact":
            adjacent_below = _first_value_below_adjacent(
                worksheet, cell.row, cell.col, merged_lookup, field
            )
            if adjacent_below:
                return adjacent_below
            continue

        below_value = _first_value_below(worksheet, cell.row, cell.col, merged_lookup, field)
        if below_value:
            return below_value
    return ""


def _extract_section(field: str, rows: dict[int, list[CellText]]) -> str:
    row_numbers = list(rows.keys())
    start_index = None

    for index, row_number in enumerate(row_numbers):
        row_text = _join_row(rows[row_number])
        if _label_match(field, row_text):
            start_index = index
            break

    if start_index is None:
        return ""

    collected: list[str] = []
    for row_number in row_numbers[start_index + 1 :]:
        row_text = _join_row(rows[row_number])
        if not row_text:
            continue
        if _is_section_header(row_text):
            break
        collected.append(row_text)

    return _normalize_multiline(" ".join(collected))


def _first_value_to_right(
    worksheet: Worksheet,
    row: int,
    label_col: int,
    merged_lookup: dict[tuple[int, int], str],
    field: str = "",
) -> str:
    seen: set[str] = set()
    for col in range(label_col + 1, worksheet.max_column + 1):
        value = _clean_value(worksheet.cell(row, col).value)
        if not value:
            value = merged_lookup.get((row, col), "")
        if not value or value in seen:
            continue
        seen.add(value)
        if _is_valid_inline_value(field, value):
            return value
    return ""


def _first_value_below(
    worksheet: Worksheet,
    label_row: int,
    col: int,
    merged_lookup: dict[tuple[int, int], str],
    field: str = "",
) -> str:
    seen: set[str] = set()
    for row in range(label_row + 1, worksheet.max_row + 1):
        value = _clean_value(worksheet.cell(row, col).value)
        if not value:
            value = merged_lookup.get((row, col), "")
        if not value or value in seen:
            continue
        seen.add(value)
        if _is_valid_inline_value(field, value):
            return value
    return ""


def _first_value_below_adjacent(
    worksheet: Worksheet,
    label_row: int,
    label_col: int,
    merged_lookup: dict[tuple[int, int], str],
    field: str,
    max_offset: int = 3,
) -> str:
    for offset in range(1, max_offset + 1):
        value = _first_value_below(
            worksheet,
            label_row,
            label_col + offset,
            merged_lookup,
            field,
        )
        if value:
            return value
    return ""


def _is_valid_inline_value(field: str, value: str) -> bool:
    if _is_any_label(value):
        return False
    if field == "Contact" and _looks_like_date(value):
        return False
    return True


def _looks_like_date(value: str) -> bool:
    return bool(re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", value.strip()))


def _join_row(cells: list[CellText]) -> str:
    return _normalize_multiline(" ".join(cell.value for cell in cells))


def _label_match(field: str, value: str) -> bool:
    normalized = _normalize_label_text(value)
    for alias in LABEL_ALIASES[field]:
        alias_text = _normalize_label_text(alias)
        pattern = rf"(^|\s){re.escape(alias_text)}($|\s)"
        if re.search(pattern, normalized):
            return True
    return False


def _value_after_label(field: str, value: str) -> str:
    for alias in sorted(LABEL_ALIASES[field], key=len, reverse=True):
        match = re.search(re.escape(alias), value, flags=re.IGNORECASE)
        if not match:
            continue
        tail = value[match.end() :]
        tail = re.sub(r"^[\s:;#.-]+", "", tail)
        return _normalize_multiline(tail)
    return ""


def _is_section_header(value: str) -> bool:
    normalized = _normalize_label_text(value)
    return any(
        re.fullmatch(rf"{re.escape(_normalize_label_text(alias))}[\s:;#.-]*", normalized)
        for field in SECTION_FIELDS
        for alias in LABEL_ALIASES[field]
    )


def _is_any_label(value: str) -> bool:
    normalized = _normalize_label_text(value)
    return any(
        re.fullmatch(rf"{re.escape(_normalize_label_text(alias))}[\s:;#.-]*", normalized)
        for aliases in LABEL_ALIASES.values()
        for alias in aliases
    )


def _normalize_label_text(value: str) -> str:
    text = re.sub(r"[^A-Z0-9#&]+", " ", value.upper())
    return re.sub(r"\s+", " ", text).strip()


def _normalize_multiline(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def _clean_value(value: Any) -> str:
    if value is None:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y")
    return _normalize_multiline(str(value))
