from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .extractor import ExtractionResult
from .fields import FIELD_COLUMNS


def read_existing_rows(existing_path: str | Path) -> list[dict[str, str]]:
    workbook = load_workbook(existing_path, data_only=True, read_only=True)
    worksheet = workbook.active

    rows = worksheet.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        workbook.close()
        return []

    column_index = {
        str(name).strip(): index
        for index, name in enumerate(header)
        if name is not None and str(name).strip()
    }

    existing_rows: list[dict[str, str]] = []
    for row in rows:
        if not row or all(value in (None, "") for value in row):
            continue
        data = {column: "" for column in FIELD_COLUMNS}
        for column in FIELD_COLUMNS:
            index = column_index.get(column)
            if index is None or index >= len(row):
                continue
            value = row[index]
            data[column] = "" if value is None else str(value).strip()
        existing_rows.append(data)

    workbook.close()
    return existing_rows


def write_results(
    results: list[ExtractionResult],
    output_path: str | Path,
    existing_path: str | Path | None = None,
) -> None:
    rows: list[dict[str, str]] = []
    if existing_path:
        rows.extend(read_existing_rows(existing_path))

    for result in results:
        rows.append(result.data)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Extracted Data"

    worksheet.append(FIELD_COLUMNS)
    for row in rows:
        worksheet.append([row.get(column, "") for column in FIELD_COLUMNS])

    _style_sheet(worksheet)
    workbook.save(output_path)


def _style_sheet(worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 24,
        "B": 18,
        "C": 26,
        "D": 10,
        "E": 10,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 18,
    }
    for index, column_name in enumerate(FIELD_COLUMNS, start=1):
        letter = get_column_letter(index)
        worksheet.column_dimensions[letter].width = widths.get(letter, min(max(len(column_name) + 8, 18), 45))

    worksheet.freeze_panes = "A2"
